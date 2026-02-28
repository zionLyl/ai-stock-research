#!/usr/bin/env python3
"""
Excel Model Builder
openpyxl helpers for building financial models (DCF, comps, etc.).

Usage:
    from excel_builder import ExcelModel
    model = ExcelModel("AAPL_DCF_Model.xlsx")
    model.create_dcf_template()
    model.save()
"""

import sys
from datetime import datetime
from pathlib import Path
from typing import Optional, Union

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        NamedStyle,
        PatternFill,
        Side,
        numbers,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ── Style Constants ───────────────────────────────────────────────────

# Colors
DARK_BLUE = "17365D"
MED_BLUE = "4472C4"
LIGHT_BLUE = "D9E2F3"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
INPUT_BLUE = "0000FF"   # Blue font for hard-coded inputs
BLACK = "000000"
GREEN = "008000"
RED = "CC0000"
ORANGE = "FF8C00"

# Fonts
FONT_HEADER = Font(name="Calibri", size=14, bold=True, color=WHITE)
FONT_SECTION = Font(name="Calibri", size=12, bold=True, color=WHITE)
FONT_COL_HEADER = Font(name="Calibri", size=10, bold=True, color=BLACK)
FONT_DATA = Font(name="Calibri", size=10, color=BLACK)
FONT_INPUT = Font(name="Calibri", size=10, color=INPUT_BLUE)  # Inputs in blue
FONT_FORMULA = Font(name="Calibri", size=10, color=BLACK)     # Formulas in black
FONT_LABEL = Font(name="Calibri", size=10, bold=False, color=BLACK)
FONT_BOLD = Font(name="Calibri", size=10, bold=True, color=BLACK)
FONT_SMALL = Font(name="Calibri", size=9, color="666666")
FONT_STAT = Font(name="Calibri", size=10, italic=True, color=BLACK)

# Fills
FILL_HEADER = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
FILL_SECTION = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
FILL_COL_HEADER = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
FILL_STATS = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_INPUT = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")  # Light yellow for inputs
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

# Borders
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D0D0D0"),
)
BOTTOM_BORDER = Border(
    bottom=Side(style="medium", color=BLACK),
)
DOUBLE_BOTTOM = Border(
    bottom=Side(style="double", color=BLACK),
)

# Alignments
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Number Formats
FMT_NUMBER = '#,##0'
FMT_DECIMAL1 = '#,##0.0'
FMT_DECIMAL2 = '#,##0.00'
FMT_PERCENT = '0.0%'
FMT_PERCENT2 = '0.00%'
FMT_MULTIPLE = '#,##0.0"x"'
FMT_MONEY = '$#,##0'
FMT_MONEY_M = '$#,##0.0,,"M"'
FMT_MONEY_B = '$#,##0.0,,,"B"'
FMT_DATE = 'YYYY-MM-DD'


class ExcelModel:
    """Build financial models in Excel using openpyxl."""

    def __init__(self, filename: str, output_dir: Optional[str] = None):
        self.wb = Workbook()
        self.filename = filename
        if output_dir:
            self.output_path = Path(output_dir) / filename
        else:
            reports_dir = Path.home() / ".openclaw" / "workspace" / "reports"
            reports_dir.mkdir(parents=True, exist_ok=True)
            self.output_path = reports_dir / filename
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

    def save(self) -> str:
        """Save workbook and return the file path."""
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(self.output_path))
        return str(self.output_path)

    # ── Sheet Management ──────────────────────────────────────────────

    def add_sheet(self, name: str, tab_color: Optional[str] = None):
        """Add a new worksheet."""
        ws = self.wb.create_sheet(title=name)
        if tab_color:
            ws.sheet_properties.tabColor = tab_color
        return ws

    def get_sheet(self, name: str):
        """Get existing worksheet by name."""
        return self.wb[name]

    # ── Formatting Helpers ────────────────────────────────────────────

    @staticmethod
    def set_column_widths(ws, widths: dict):
        """Set column widths. widths = {1: 25, 2: 15, ...} (1-indexed)."""
        for col_num, width in widths.items():
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = width

    @staticmethod
    def write_header_row(ws, row: int, texts: list, start_col: int = 1):
        """Write a dark blue header row spanning multiple columns."""
        for i, text in enumerate(texts):
            cell = ws.cell(row=row, column=start_col + i, value=text)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER

    @staticmethod
    def write_section_header(ws, row: int, text: str, col_span: int = 10, start_col: int = 1):
        """Write a medium blue section header row."""
        for c in range(start_col, start_col + col_span):
            cell = ws.cell(row=row, column=c)
            cell.fill = FILL_SECTION
        cell = ws.cell(row=row, column=start_col, value=text)
        cell.font = FONT_SECTION
        cell.fill = FILL_SECTION
        cell.alignment = ALIGN_LEFT

    @staticmethod
    def write_column_headers(ws, row: int, headers: list, start_col: int = 1):
        """Write column headers with light blue background."""
        for i, header in enumerate(headers):
            cell = ws.cell(row=row, column=start_col + i, value=header)
            cell.font = FONT_COL_HEADER
            cell.fill = FILL_COL_HEADER
            cell.alignment = ALIGN_CENTER

    @staticmethod
    def write_data_row(
        ws,
        row: int,
        values: list,
        start_col: int = 1,
        is_input: bool = False,
        fmt: Optional[str] = None,
        bold: bool = False,
    ):
        """Write a row of data values."""
        for i, val in enumerate(values):
            cell = ws.cell(row=row, column=start_col + i, value=val)
            if i == 0:
                # First column is usually a label
                cell.font = FONT_BOLD if bold else FONT_LABEL
                cell.alignment = ALIGN_LEFT
            else:
                cell.font = FONT_INPUT if is_input else FONT_FORMULA
                cell.alignment = ALIGN_CENTER
                if is_input:
                    cell.fill = FILL_INPUT
                if fmt:
                    cell.number_format = fmt

    @staticmethod
    def write_stats_row(ws, row: int, values: list, start_col: int = 1, fmt: Optional[str] = None):
        """Write a statistics row (gray background, italic)."""
        for i, val in enumerate(values):
            cell = ws.cell(row=row, column=start_col + i, value=val)
            cell.font = FONT_STAT
            cell.fill = FILL_STATS
            cell.alignment = ALIGN_CENTER if i > 0 else ALIGN_LEFT
            if fmt and i > 0:
                cell.number_format = fmt

    @staticmethod
    def add_cell_comment(ws, row: int, col: int, comment_text: str):
        """Add a comment to a cell (for source attribution)."""
        from openpyxl.comments import Comment

        cell = ws.cell(row=row, column=col)
        cell.comment = Comment(comment_text, "OpenClaw Research")

    @staticmethod
    def write_formula(ws, row: int, col: int, formula: str, fmt: Optional[str] = None):
        """Write a formula to a cell."""
        cell = ws.cell(row=row, column=col, value=formula)
        cell.font = FONT_FORMULA
        cell.alignment = ALIGN_CENTER
        if fmt:
            cell.number_format = fmt

    @staticmethod
    def apply_border(ws, row: int, col_start: int, col_end: int, border_style=BOTTOM_BORDER):
        """Apply border to a range of cells in a row."""
        for c in range(col_start, col_end + 1):
            ws.cell(row=row, column=c).border = border_style

    # ── Sensitivity Table ─────────────────────────────────────────────

    def create_sensitivity_table(
        self,
        ws,
        start_row: int,
        start_col: int,
        title: str,
        row_label: str,
        col_label: str,
        row_values: list,
        col_values: list,
        base_formula_template: str,
        fmt: str = FMT_MONEY,
    ):
        """
        Create a sensitivity analysis table.
        base_formula_template: Python format string with {row_val} and {col_val} placeholders.
        """
        # Title
        cell = ws.cell(row=start_row, column=start_col, value=title)
        cell.font = FONT_BOLD
        cell.alignment = ALIGN_LEFT

        # Column label
        cell = ws.cell(row=start_row + 1, column=start_col, value=f"{row_label} \\ {col_label}")
        cell.font = FONT_COL_HEADER
        cell.fill = FILL_COL_HEADER
        cell.alignment = ALIGN_CENTER

        # Column headers (col_values)
        for j, cv in enumerate(col_values):
            cell = ws.cell(row=start_row + 1, column=start_col + 1 + j, value=cv)
            cell.font = FONT_COL_HEADER
            cell.fill = FILL_COL_HEADER
            cell.alignment = ALIGN_CENTER
            if isinstance(cv, float) and cv < 1:
                cell.number_format = FMT_PERCENT

        # Row headers and formulas
        for i, rv in enumerate(row_values):
            cell = ws.cell(row=start_row + 2 + i, column=start_col, value=rv)
            cell.font = FONT_LABEL
            cell.fill = FILL_STATS
            cell.alignment = ALIGN_CENTER
            if isinstance(rv, float) and rv < 1:
                cell.number_format = FMT_PERCENT

            for j, cv in enumerate(col_values):
                cell = ws.cell(
                    row=start_row + 2 + i,
                    column=start_col + 1 + j,
                )
                # Each cell gets the formula template populated
                try:
                    formula = base_formula_template.format(row_val=rv, col_val=cv)
                    cell.value = formula
                except (KeyError, IndexError):
                    cell.value = ""
                cell.font = FONT_DATA
                cell.alignment = ALIGN_CENTER
                cell.number_format = fmt

        return start_row + 2 + len(row_values)

    # ── Chart Helpers ─────────────────────────────────────────────────

    @staticmethod
    def add_bar_chart(
        ws,
        title: str,
        data_range: tuple,
        category_range: tuple,
        position: str = "A1",
        width: int = 18,
        height: int = 12,
    ):
        """
        Add a bar chart to the worksheet.
        data_range: (min_col, min_row, max_col, max_row)
        category_range: (min_col, min_row, max_col, max_row)
        """
        chart = BarChart()
        chart.type = "col"
        chart.title = title
        chart.width = width
        chart.height = height

        data = Reference(ws, *data_range)
        categories = Reference(ws, *category_range)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, position)
        return chart

    @staticmethod
    def add_line_chart(
        ws,
        title: str,
        data_range: tuple,
        category_range: tuple,
        position: str = "A1",
        width: int = 18,
        height: int = 12,
    ):
        """Add a line chart to the worksheet."""
        chart = LineChart()
        chart.title = title
        chart.width = width
        chart.height = height

        data = Reference(ws, *data_range)
        categories = Reference(ws, *category_range)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, position)
        return chart

    # ── DCF Template ──────────────────────────────────────────────────

    def create_dcf_template(
        self,
        company: str,
        years: int = 5,
        historical_years: int = 3,
    ) -> str:
        """
        Create a complete DCF model template.
        Returns the file path of the saved workbook.
        """
        total_cols = 2 + historical_years + years  # Label + Historical + Projected

        # ── Assumptions Sheet ──
        ws_assumptions = self.add_sheet("Assumptions", tab_color="4472C4")
        self._build_dcf_assumptions(ws_assumptions, company, years, total_cols)

        # ── Income Statement Sheet ──
        ws_income = self.add_sheet("Income Statement", tab_color="548235")
        self._build_income_statement(ws_income, company, years, historical_years, total_cols)

        # ── Free Cash Flow Sheet ──
        ws_fcf = self.add_sheet("Free Cash Flow", tab_color="BF8F00")
        self._build_fcf_sheet(ws_fcf, company, years, historical_years, total_cols)

        # ── DCF Valuation Sheet ──
        ws_dcf = self.add_sheet("DCF Valuation", tab_color="C00000")
        self._build_dcf_valuation(ws_dcf, company, years, total_cols)

        # ── Sensitivity Analysis Sheet ──
        ws_sens = self.add_sheet("Sensitivity", tab_color="7030A0")
        self._build_sensitivity(ws_sens, company)

        return self.save()

    def _build_dcf_assumptions(self, ws, company: str, years: int, total_cols: int):
        """Build the Assumptions tab."""
        self.write_header_row(ws, 1, [f"{company} — DCF Assumptions"], start_col=1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        self.set_column_widths(ws, {1: 35, 2: 15, 3: 15, 4: 15, 5: 15, 6: 15})

        row = 3
        self.write_section_header(ws, row, "Revenue Assumptions", col_span=6)
        row += 1
        self.write_column_headers(ws, row, ["Metric", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"])
        row += 1
        self.write_data_row(ws, row, ["Revenue Growth Rate (%)", "", "", "", "", ""], is_input=True, fmt=FMT_PERCENT)
        row += 1
        self.write_data_row(ws, row, ["Revenue ($M)", "", "", "", "", ""], fmt=FMT_NUMBER)
        row += 2

        self.write_section_header(ws, row, "Margin Assumptions", col_span=6)
        row += 1
        self.write_column_headers(ws, row, ["Metric", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"])
        row += 1
        for label in ["Gross Margin (%)", "Operating Margin (%)", "Tax Rate (%)", "D&A (% of Revenue)", "CapEx (% of Revenue)", "NWC Change (% of Revenue)"]:
            self.write_data_row(ws, row, [label, "", "", "", "", ""], is_input=True, fmt=FMT_PERCENT)
            row += 1
        row += 1

        self.write_section_header(ws, row, "WACC Assumptions", col_span=6)
        row += 1
        wacc_items = [
            "Risk-Free Rate (%)",
            "Equity Risk Premium (%)",
            "Beta (Levered)",
            "Cost of Equity (%)",
            "Pre-Tax Cost of Debt (%)",
            "Tax Rate (%)",
            "After-Tax Cost of Debt (%)",
            "Debt / Total Capital (%)",
            "Equity / Total Capital (%)",
            "WACC (%)",
        ]
        for label in wacc_items:
            self.write_data_row(ws, row, [label, ""], is_input=True, fmt=FMT_PERCENT)
            row += 1
        row += 1

        self.write_section_header(ws, row, "Terminal Value Assumptions", col_span=6)
        row += 1
        for label in ["Terminal Growth Rate (%)", "Exit EV/EBITDA Multiple (x)"]:
            self.write_data_row(ws, row, [label, ""], is_input=True)
            row += 1

    def _build_income_statement(self, ws, company: str, years: int, hist_years: int, total_cols: int):
        """Build the Income Statement tab."""
        self.write_header_row(ws, 1, [f"{company} — Income Statement ($M)"], start_col=1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        col_widths = {1: 30}
        for i in range(2, total_cols + 1):
            col_widths[i] = 14
        self.set_column_widths(ws, col_widths)

        # Column headers
        headers = [""]
        current_year = datetime.now().year
        for y in range(hist_years, 0, -1):
            headers.append(f"FY{current_year - y}A")
        for y in range(1, years + 1):
            headers.append(f"FY{current_year + y - 1}E")
        self.write_column_headers(ws, 3, headers)

        # Line items
        row = 5
        income_items = [
            ("Revenue", FMT_NUMBER, True),
            ("  Revenue Growth (%)", FMT_PERCENT, False),
            ("Cost of Goods Sold", FMT_NUMBER, False),
            ("Gross Profit", FMT_NUMBER, True),
            ("  Gross Margin (%)", FMT_PERCENT, False),
            ("", None, False),
            ("Operating Expenses:", None, False),
            ("  R&D", FMT_NUMBER, False),
            ("  SG&A", FMT_NUMBER, False),
            ("  Other Operating", FMT_NUMBER, False),
            ("Total Operating Expenses", FMT_NUMBER, True),
            ("", None, False),
            ("Operating Income (EBIT)", FMT_NUMBER, True),
            ("  Operating Margin (%)", FMT_PERCENT, False),
            ("", None, False),
            ("Interest Expense", FMT_NUMBER, False),
            ("Other Income / (Expense)", FMT_NUMBER, False),
            ("Pre-Tax Income", FMT_NUMBER, True),
            ("Income Tax", FMT_NUMBER, False),
            ("  Effective Tax Rate (%)", FMT_PERCENT, False),
            ("Net Income", FMT_NUMBER, True),
            ("  Net Margin (%)", FMT_PERCENT, False),
            ("", None, False),
            ("EBITDA", FMT_NUMBER, True),
            ("  EBITDA Margin (%)", FMT_PERCENT, False),
            ("", None, False),
            ("Shares Outstanding (M)", FMT_DECIMAL1, False),
            ("EPS", FMT_DECIMAL2, False),
        ]
        for label, fmt, bold in income_items:
            values = [label] + [""] * (total_cols - 1)
            self.write_data_row(ws, row, values, fmt=fmt, bold=bold)
            if bold:
                self.apply_border(ws, row, 1, total_cols, THIN_BORDER)
            row += 1

    def _build_fcf_sheet(self, ws, company: str, years: int, hist_years: int, total_cols: int):
        """Build the Free Cash Flow tab."""
        self.write_header_row(ws, 1, [f"{company} — Free Cash Flow ($M)"], start_col=1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        col_widths = {1: 30}
        for i in range(2, total_cols + 1):
            col_widths[i] = 14
        self.set_column_widths(ws, col_widths)

        headers = [""]
        current_year = datetime.now().year
        for y in range(1, years + 1):
            headers.append(f"FY{current_year + y - 1}E")
        self.write_column_headers(ws, 3, headers[:years + 1])

        row = 5
        fcf_items = [
            ("EBIT", FMT_NUMBER, False),
            ("  Less: Taxes on EBIT", FMT_NUMBER, False),
            ("NOPAT (Net Operating Profit After Tax)", FMT_NUMBER, True),
            ("", None, False),
            ("  Plus: Depreciation & Amortization", FMT_NUMBER, False),
            ("  Less: Capital Expenditures", FMT_NUMBER, False),
            ("  Less: Change in Net Working Capital", FMT_NUMBER, False),
            ("", None, False),
            ("Unlevered Free Cash Flow (UFCF)", FMT_NUMBER, True),
            ("", None, False),
            ("Discount Factor", FMT_DECIMAL2, False),
            ("Present Value of UFCF", FMT_NUMBER, True),
        ]
        for label, fmt, bold in fcf_items:
            values = [label] + [""] * years
            self.write_data_row(ws, row, values, fmt=fmt, bold=bold)
            if bold:
                self.apply_border(ws, row, 1, years + 1, THIN_BORDER)
            row += 1

    def _build_dcf_valuation(self, ws, company: str, years: int, total_cols: int):
        """Build the DCF Valuation summary tab."""
        self.write_header_row(ws, 1, [f"{company} — DCF Valuation Summary"], start_col=1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        self.set_column_widths(ws, {1: 35, 2: 18, 3: 18, 4: 18})

        row = 3
        self.write_section_header(ws, row, "Enterprise Value Bridge", col_span=4)
        row += 1
        self.write_column_headers(ws, row, ["Component", "Value ($M)", "Notes", ""])
        row += 1
        bridge_items = [
            "PV of Free Cash Flows",
            "PV of Terminal Value (Gordon Growth)",
            "PV of Terminal Value (Exit Multiple)",
            "",
            "Enterprise Value (Gordon Growth)",
            "Enterprise Value (Exit Multiple)",
            "Enterprise Value (Blended Average)",
        ]
        for item in bridge_items:
            bold = item.startswith("Enterprise Value")
            self.write_data_row(ws, row, [item, "", ""], fmt=FMT_NUMBER, bold=bold)
            if bold:
                self.apply_border(ws, row, 1, 4, THIN_BORDER)
            row += 1

        row += 1
        self.write_section_header(ws, row, "Equity Value Bridge", col_span=4)
        row += 1
        equity_items = [
            "Enterprise Value",
            "  Less: Total Debt",
            "  Plus: Cash & Equivalents",
            "  Less: Minority Interest",
            "  Less: Preferred Equity",
            "",
            "Equity Value",
            "Shares Outstanding (M)",
            "Implied Share Price",
            "",
            "Current Share Price",
            "Upside / (Downside) %",
        ]
        for item in equity_items:
            bold = item in ("Equity Value", "Implied Share Price", "Upside / (Downside) %")
            fmt = FMT_PERCENT if "%" in item and "Upside" in item else FMT_NUMBER
            if "Share Price" in item:
                fmt = FMT_DECIMAL2
            self.write_data_row(ws, row, [item, "", ""], fmt=fmt, bold=bold)
            if bold:
                self.apply_border(ws, row, 1, 4, THIN_BORDER)
            row += 1

    def _build_sensitivity(self, ws, company: str):
        """Build the Sensitivity Analysis tab."""
        self.write_header_row(ws, 1, [f"{company} — Sensitivity Analysis"], start_col=1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        self.set_column_widths(ws, {i: 14 for i in range(1, 9)})

        # Table 1: WACC vs Terminal Growth Rate → Implied Share Price
        row = 3
        cell = ws.cell(row=row, column=1, value="Table 1: WACC vs. Terminal Growth Rate → Implied Share Price")
        cell.font = FONT_BOLD

        row += 1
        wacc_values = [0.08, 0.09, 0.10, 0.11, 0.12]
        tgr_values = [0.015, 0.020, 0.025, 0.030, 0.035]

        cell = ws.cell(row=row, column=1, value="WACC \\ TGR")
        cell.font = FONT_COL_HEADER
        cell.fill = FILL_COL_HEADER
        for j, tgr in enumerate(tgr_values):
            cell = ws.cell(row=row, column=2 + j, value=tgr)
            cell.font = FONT_COL_HEADER
            cell.fill = FILL_COL_HEADER
            cell.number_format = FMT_PERCENT

        for i, wacc in enumerate(wacc_values):
            cell = ws.cell(row=row + 1 + i, column=1, value=wacc)
            cell.font = FONT_LABEL
            cell.fill = FILL_STATS
            cell.number_format = FMT_PERCENT
            for j in range(len(tgr_values)):
                cell = ws.cell(row=row + 1 + i, column=2 + j, value="")
                cell.font = FONT_DATA
                cell.alignment = ALIGN_CENTER
                cell.number_format = FMT_DECIMAL2

        # Table 2: WACC vs Exit Multiple → Implied Share Price
        row += len(wacc_values) + 3
        cell = ws.cell(row=row, column=1, value="Table 2: WACC vs. Exit EV/EBITDA Multiple → Implied Share Price")
        cell.font = FONT_BOLD

        row += 1
        exit_multiples = [8.0, 10.0, 12.0, 14.0, 16.0]

        cell = ws.cell(row=row, column=1, value="WACC \\ Multiple")
        cell.font = FONT_COL_HEADER
        cell.fill = FILL_COL_HEADER
        for j, mult in enumerate(exit_multiples):
            cell = ws.cell(row=row, column=2 + j, value=mult)
            cell.font = FONT_COL_HEADER
            cell.fill = FILL_COL_HEADER
            cell.number_format = FMT_MULTIPLE

        for i, wacc in enumerate(wacc_values):
            cell = ws.cell(row=row + 1 + i, column=1, value=wacc)
            cell.font = FONT_LABEL
            cell.fill = FILL_STATS
            cell.number_format = FMT_PERCENT
            for j in range(len(exit_multiples)):
                cell = ws.cell(row=row + 1 + i, column=2 + j, value="")
                cell.font = FONT_DATA
                cell.alignment = ALIGN_CENTER
                cell.number_format = FMT_DECIMAL2

        # Table 3: Revenue Growth vs Operating Margin → Implied Share Price
        row += len(wacc_values) + 3
        cell = ws.cell(row=row, column=1, value="Table 3: Revenue Growth vs. Operating Margin → Implied Share Price")
        cell.font = FONT_BOLD

        row += 1
        rev_growth = [0.03, 0.05, 0.08, 0.10, 0.12]
        op_margins = [0.15, 0.18, 0.20, 0.22, 0.25]

        cell = ws.cell(row=row, column=1, value="Growth \\ Margin")
        cell.font = FONT_COL_HEADER
        cell.fill = FILL_COL_HEADER
        for j, margin in enumerate(op_margins):
            cell = ws.cell(row=row, column=2 + j, value=margin)
            cell.font = FONT_COL_HEADER
            cell.fill = FILL_COL_HEADER
            cell.number_format = FMT_PERCENT

        for i, growth in enumerate(rev_growth):
            cell = ws.cell(row=row + 1 + i, column=1, value=growth)
            cell.font = FONT_LABEL
            cell.fill = FILL_STATS
            cell.number_format = FMT_PERCENT
            for j in range(len(op_margins)):
                cell = ws.cell(row=row + 1 + i, column=2 + j, value="")
                cell.font = FONT_DATA
                cell.alignment = ALIGN_CENTER
                cell.number_format = FMT_DECIMAL2

    # ── Comps Template ────────────────────────────────────────────────

    def create_comps_template(
        self,
        company: str,
        peers: list[str],
        metrics: Optional[list[str]] = None,
    ) -> str:
        """
        Create a comparable company analysis template.
        Returns the file path of the saved workbook.
        """
        all_companies = [company] + peers
        num_companies = len(all_companies)

        # ── Operating Metrics Sheet ──
        ws_ops = self.add_sheet("Operating Metrics", tab_color="548235")
        self._build_comps_operating(ws_ops, company, all_companies, metrics)

        # ── Valuation Multiples Sheet ──
        ws_val = self.add_sheet("Valuation Multiples", tab_color="4472C4")
        self._build_comps_valuation(ws_val, company, all_companies)

        # ── Notes Sheet ──
        ws_notes = self.add_sheet("Notes & Methodology", tab_color="7030A0")
        self._build_comps_notes(ws_notes, company)

        return self.save()

    def _build_comps_operating(self, ws, company: str, companies: list, metrics: Optional[list] = None):
        """Build Operating Metrics tab for comps."""
        n = len(companies)
        total_cols = 7  # Company + 6 metrics

        self.write_header_row(ws, 1, [f"COMPARABLE COMPANY ANALYSIS — OPERATING METRICS"], start_col=1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

        cell = ws.cell(row=2, column=1, value=f"As of [Date] | All figures in USD Millions")
        cell.font = FONT_SMALL

        self.set_column_widths(ws, {1: 22, 2: 16, 3: 14, 4: 16, 5: 14, 6: 16, 7: 14})

        headers = ["Company", "Revenue (LTM)", "Rev Growth %", "Gross Profit", "Gross Margin", "EBITDA", "EBITDA Margin"]
        self.write_column_headers(ws, 4, headers)

        row = 5
        for company_name in companies:
            self.write_data_row(ws, row, [company_name, "", "", "", "", "", ""])
            row += 1

        # Blank row
        row += 1

        # Statistics
        stats_labels = ["Max", "75th Percentile", "Median", "25th Percentile", "Minimum"]
        for label in stats_labels:
            self.write_stats_row(ws, row, [label, "", "", "", "", "", ""])
            row += 1

    def _build_comps_valuation(self, ws, company: str, companies: list):
        """Build Valuation Multiples tab for comps."""
        n = len(companies)
        total_cols = 7

        self.write_header_row(ws, 1, [f"COMPARABLE COMPANY ANALYSIS — VALUATION MULTIPLES"], start_col=1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

        cell = ws.cell(row=2, column=1, value=f"As of [Date] | All figures in USD Millions")
        cell.font = FONT_SMALL

        self.set_column_widths(ws, {1: 22, 2: 16, 3: 16, 4: 14, 5: 14, 6: 14, 7: 14})

        headers = ["Company", "Market Cap", "Enterprise Value", "EV/Revenue", "EV/EBITDA", "P/E", "FCF Yield"]
        self.write_column_headers(ws, 4, headers)

        row = 5
        for company_name in companies:
            self.write_data_row(ws, row, [company_name, "", "", "", "", "", ""])
            row += 1

        row += 1
        stats_labels = ["Max", "75th Percentile", "Median", "25th Percentile", "Minimum"]
        for label in stats_labels:
            self.write_stats_row(ws, row, [label, "", "", "", "", "", ""])
            row += 1

    def _build_comps_notes(self, ws, company: str):
        """Build Notes & Methodology tab."""
        self.set_column_widths(ws, {1: 25, 2: 60})

        self.write_header_row(ws, 1, [f"Notes & Methodology"], start_col=1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

        row = 3
        notes_items = [
            ("Data Sources", "[List all data sources: yfinance, SEC EDGAR, etc.]"),
            ("Time Period", "[LTM / Fiscal Year / Calendar Year]"),
            ("Currency", "USD"),
            ("Units", "Millions unless otherwise noted"),
            ("", ""),
            ("Definitions", ""),
            ("Revenue (LTM)", "Last twelve months total revenue"),
            ("EBITDA", "Operating Income + Depreciation & Amortization"),
            ("Enterprise Value", "Market Cap + Total Debt - Cash & Equivalents"),
            ("FCF Yield", "Free Cash Flow / Market Cap"),
            ("", ""),
            ("Methodology Notes", ""),
            ("Peer Selection", "[Rationale for peer group selection]"),
            ("Adjustments", "[Any one-time or non-recurring adjustments]"),
            ("Limitations", "[Data limitations or caveats]"),
        ]
        for label, value in notes_items:
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = FONT_BOLD if not label.startswith(" ") else FONT_LABEL
            cell = ws.cell(row=row, column=2, value=value)
            cell.font = FONT_DATA
            cell.alignment = ALIGN_WRAP
            row += 1

    # ── Portfolio Tracker Template ────────────────────────────────────

    def create_portfolio_template(self) -> str:
        """Create a portfolio tracking template."""
        ws = self.add_sheet("Portfolio", tab_color="548235")
        self.write_header_row(ws, 1, ["Portfolio Monitor"], start_col=1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

        self.set_column_widths(ws, {
            1: 10, 2: 20, 3: 12, 4: 12, 5: 12,
            6: 14, 7: 14, 8: 14, 9: 12, 10: 12,
        })

        headers = [
            "Ticker", "Company", "Shares", "Avg Cost", "Current Price",
            "Market Value", "Cost Basis", "P&L ($)", "P&L (%)", "Weight (%)",
        ]
        self.write_column_headers(ws, 3, headers)

        # Empty rows for data
        for row in range(4, 19):
            self.write_data_row(ws, row, [""] * 10)

        # Summary row
        row = 20
        self.apply_border(ws, row - 1, 1, 10, BOTTOM_BORDER)
        self.write_data_row(ws, row, ["", "TOTAL", "", "", "", "", "", "", "", "100.0%"], bold=True)

        return self.save()


# ── CLI Interface ─────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python excel_builder.py <TICKER> <dcf|comps|portfolio>")
        print("  python excel_builder.py AAPL dcf")
        print("  python excel_builder.py AAPL comps MSFT,GOOGL,AMZN")
        print("  python excel_builder.py - portfolio")
        sys.exit(1)

    ticker_arg = sys.argv[1].upper()
    model_type = sys.argv[2].lower()

    if model_type == "dcf":
        model = ExcelModel(f"{ticker_arg}_DCF_Model.xlsx")
        path = model.create_dcf_template(ticker_arg)
        print(f"DCF template created: {path}")

    elif model_type == "comps":
        peers = []
        if len(sys.argv) > 3:
            peers = sys.argv[3].upper().split(",")
        model = ExcelModel(f"{ticker_arg}_Comps.xlsx")
        path = model.create_comps_template(ticker_arg, peers)
        print(f"Comps template created: {path}")

    elif model_type == "portfolio":
        model = ExcelModel("Portfolio_Monitor.xlsx")
        path = model.create_portfolio_template()
        print(f"Portfolio template created: {path}")

    else:
        print(f"Unknown model type: {model_type}")
        print("Supported types: dcf, comps, portfolio")
        sys.exit(1)
